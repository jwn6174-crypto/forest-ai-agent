"""
forest_household_parse.py
가이드 §3.1 — 임가경제 데이터 (임업 다각화 보조 수입).

출처: 산림청 임가경제조사 (산림임업통계플랫폼, KOSIS 연계)
  - 임업손익계산서(충청북도) 2020~2024
  - 주요지표(월별지표) 2020~2024
  - 단위: 천원, 충북 표본임가 평균

산출물: data/processed/forest_household_economy.json
용도: 모듈 C Faustmann NPV 의 '임업 다각화 보조 수입' 입력.
      산주가 원목·탄소 외 임업으로 버는 연간 소득.
"""

import json
import openpyxl
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
RAW = ROOT / "module_bd" / "data" / "raw" / "forest_household_economy"
OUT = ROOT / "module_bd" / "data" / "processed"
OUT.mkdir(parents=True, exist_ok=True)

PNL_FILE = RAW / "임업손익계산서_충청북도_2020-2024.xlsx"
IND_FILE = RAW / "주요지표_월별지표_2020-2024.xlsx"

YEARS = [2020, 2021, 2022, 2023, 2024]


def _num(v):
    """셀 값 → float. '-'·빈칸·None 은 None."""
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "-"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_indicators() -> dict:
    """주요지표(월별) — 'year' 행에서 연도별 임가소득/임업소득/경영비.

    구조: r6 = '년' 합계 행. 연도 블록이 가로로 5개 반복.
    각 블록 시작열(2,22,42,62,82) 기준 오프셋:
      +0 임가소득, +2 임업소득, +3 임업총수입, +4 임업경영비
    """
    wb = openpyxl.load_workbook(IND_FILE, data_only=True)
    ws = wb["데이터"]
    row6 = [c.value for c in ws[7]]  # openpyxl 1-indexed → r6 = 7번째

    result = {}
    block_starts = [2, 21, 40, 59, 78]  # 0-indexed 열 위치 (19칸 간격)
    for year, start in zip(YEARS, block_starts):
        result[year] = {
            "임가소득": _num(row6[start + 0]),
            "임업소득": _num(row6[start + 2]),
            "임업총수입": _num(row6[start + 3]),
            "임업경영비": _num(row6[start + 4]),
        }
    wb.close()
    return result


def parse_pnl() -> dict:
    """임업손익계산서(충북) — 연도별 임업수입·경영비·임업소득.

    구조: r0 = 연도 헤더(F~J열 = 2020~2024).
      r2  비용/순익 소계,  r3 유동비,  r26 고정비,
      r32 당기순익(임업소득),  r34 임업수입,  r38 단기소득임산물수입
    """
    wb = openpyxl.load_workbook(PNL_FILE, data_only=True)
    ws = wb["데이터"]
    rows = list(ws.iter_rows(values_only=True))

    # 행 라벨(5단계 중 가장 깊은 라벨)로 찾기 위한 헬퍼
    def row_by_label(*labels):
        """라벨이 한 행의 앞 5칸 어딘가에 있으면 그 행 반환."""
        for r in rows:
            head = [str(c).strip() if c else "" for c in r[:5]]
            if any(lbl in head for lbl in labels):
                return r
        return None

    # 연도는 6~10번째 열 (index 5~9)
    out = {}
    label_map = {
        "임업수입": ("임업수입",),
        "임업소득": ("당기순익(임업소득)",),
        "유동비": ("유동비",),
        "고정비": ("고정비",),
        "단기소득임산물수입": ("단기소득 임산물 수입",),
    }
    raw_vals = {}
    for key, labels in label_map.items():
        r = row_by_label(*labels)
        raw_vals[key] = [_num(r[5 + i]) for i in range(5)] if r else [None] * 5

    for i, year in enumerate(YEARS):
        flow = raw_vals["유동비"][i]
        fixed = raw_vals["고정비"][i]
        expense = None
        if flow is not None and fixed is not None:
            expense = round(flow + fixed, 1)
        out[year] = {
            "임업수입": raw_vals["임업수입"][i],
            "임업경영비": expense,            # 유동비 + 고정비
            "임업소득": raw_vals["임업소득"][i],  # 당기순익
            "단기소득임산물수입": raw_vals["단기소득임산물수입"][i],
        }
    wb.close()
    return out


def _avg(values):
    """None 제외 평균."""
    nums = [v for v in values if v is not None]
    return round(sum(nums) / len(nums), 1) if nums else None


def build_json() -> dict:
    ind = parse_indicators()
    pnl = parse_pnl()

    # 5년 평균 (임가경제 baseline)
    avg = {
        "임가소득": _avg([ind[y]["임가소득"] for y in YEARS]),
        "임업소득": _avg([ind[y]["임업소득"] for y in YEARS]),
        "임업총수입": _avg([ind[y]["임업총수입"] for y in YEARS]),
        "임업경영비": _avg([ind[y]["임업경영비"] for y in YEARS]),
    }

    data = {
        "source": {
            "survey": "산림청 임가경제조사",
            "platform": "산림임업통계플랫폼 (KOSIS 연계)",
            "region": "충청북도",
            "tables": ["임업손익계산서(충청북도)", "주요지표(월별지표)"],
            "period": "2020~2024",
            "unit": "천원",
            "note": "충북 표본임가 연 평균. Faustmann NPV 임업 다각화 보조 수입용.",
            "정의_차이_주의": (
                "두 출처의 '임업소득'은 정의가 다름. "
                "주요지표 = 임업총수입 − 임업경영비 (가구 단위, 단순). "
                "손익계산서 = 수익 − 비용(감가상각 등 포함, 회계 손익). "
                "Faustmann 보조 수입은 by_year_indicators(주요지표) 사용 권장."
            ),
        },
        "by_year_indicators": ind,   # 주요지표 출처
        "by_year_pnl": pnl,          # 손익계산서 출처
        "five_year_average": avg,    # 5년 baseline
    }

    out_path = OUT / "forest_household_economy.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"💾 {out_path.relative_to(ROOT)}")
    return data


if __name__ == "__main__":
    print("=" * 60)
    print("🌲 임가경제 데이터 파싱 — 충북 2020~2024")
    print("=" * 60)

    data = build_json()

    print("\n[주요지표 출처] 연도별 (천원):")
    for y in YEARS:
        d = data["by_year_indicators"][y]
        print(f"   {y}: 임가소득 {d['임가소득']:,} / "
              f"임업소득 {d['임업소득']:,} / 경영비 {d['임업경영비']:,}")

    print("\n[손익계산서 출처] 연도별 (천원):")
    for y in YEARS:
        d = data["by_year_pnl"][y]
        inc = d["임업소득"]
        exp = d["임업경영비"]
        print(f"   {y}: 임업수입 {d['임업수입']} / "
              f"경영비 {exp} / 임업소득 {inc}")

    print("\n[5년 평균 baseline] (천원):")
    for k, v in data["five_year_average"].items():
        print(f"   {k}: {v:,}")

    print("\n" + "=" * 60)
    print("✅ forest_household_economy.json 생성 완료")
    print("=" * 60)