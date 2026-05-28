"""
nfi5_extract.py — NFI 5차 충북 추출 (D16 시계열 확장).

진단 (nfi5_coord_probe.py + 수고 진단) 결과 — 5차 통합 장벽 모두 해소:
  · 좌표: 6/7차와 동일 범위 (EPSG:5181, 변환 불필요)
  · ID: 1078 공통 고정표본점 (형식 동일)
  · 임상/영급 코드: 6/7차와 동일 ('침엽수림(D)', '5영급')
  · 측정연도: 2006-2010 (시점 3 확보)
  · 수고: cm 단위 (340~2590, 6/7차 동일) → ÷100 = m

영문 → 한글 매핑 + 수고 cm→m 변환하여 6/7차 호환 csv 생성.

출력:
  · nfi5_chungbuk_stand.csv (6/7차 호환 컬럼명)
  · nfi5_chungbuk_tree.csv  (수고_m, 추정수고_m 포함)
"""
import openpyxl
import csv
from pathlib import Path

ROOT = Path(__file__).resolve().parents[3]
NFI_DIR = ROOT / "module_bd" / "data" / "raw" / "nfi"
NFI5 = NFI_DIR / "mdb_NFI_5_수정.xlsx"
STAND_OUT = NFI_DIR / "nfi5_chungbuk_stand.csv"
TREE_OUT = NFI_DIR / "nfi5_chungbuk_tree.csv"

# 임분조사표 영문 → 한글 (6/7차 호환)
STAND_MAP = {
    'SP_ID': '표본점번호',
    'GYEAR': '조사연도',
    'NAME_SIDO': '광역시도',
    'NAME_SGG': '시군구',
    'NAME_EMD': '읍면동',
    'GB_BESSEL_TM_Y': '좌표N',
    'GB_BESSEL_TM_X': '좌표E',
    'GB_DISTANCEFROMWAY': '도로로부터의거리',
    'GL_HEIGHT': '해발고',
    'GL_SLOPE': '경사',
    'GL_CODE_ASPECT_NM': '방위',
    'GF_CODE_FORESTTYPE_NM': '임상',
    'GF_CODE_CROWNDENS_NM': '수관밀도',
    'GF_CODE_DIACLASS_NM': '경급',
    'GF_CODE_AGECLASS_NM': '영급',
    'GF_CODE_OWN_NM': '소유',
}

# 입목조사표 영문 → 한글 (6/7차 호환)
# 수고/추정수고는 cm 원본 → 추출 후 ÷100 변환
TREE_MAP = {
    'SP_ID': '표본점번호',
    'GYEAR': '조사연도',
    'KORNAME': '수종명',
    '침활구분': '침활구분',
    'STP_DBH': 'DBH_cm',            # cm 그대로 (6/7차 호환)
    'basal area': '흉고단면적',
    'STP_CODE_TCHAR': '형질급코드',
    'STREE_HEIGHT': '_수고_cm',     # cm 원본 (아래서 ÷100 → 수고_m)
    'SPTV_STVOL': '추정간재적',     # 6/7차 '추정간재적'
    'SPTV_THEIGHT': '_추정수고_cm', # cm 원본 (아래서 ÷100 → 추정수고_m)
}


def cm_to_m(val):
    """cm → m (÷100). 결측/오류는 None."""
    try:
        f = float(val)
        return round(f / 100, 2)
    except (TypeError, ValueError):
        return None


def main():
    print("=" * 75)
    print("NFI 5차 충북 추출 (영문 -> 한글 매핑 + 수고 cm->m)")
    print("=" * 75)

    wb = openpyxl.load_workbook(NFI5, read_only=True)

    # ── 1. 임분조사표 ──
    print("\n[1] 임분조사표 추출...")
    ws = wb['임분조사표']
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(header)}

    stand_cols = list(STAND_MAP.keys())
    stand_out_cols = [STAND_MAP[c] for c in stand_cols]

    stand_rows = []
    chungbuk_ids = set()
    for r in ws.iter_rows(min_row=2, values_only=True):
        sido = r[idx['NAME_SIDO']]
        if sido and '충청북도' in str(sido):
            row_out = {}
            for eng in stand_cols:
                val = r[idx[eng]] if eng in idx else None
                row_out[STAND_MAP[eng]] = val
            stand_rows.append(row_out)
            chungbuk_ids.add(str(r[idx['SP_ID']]))

    print(f"  충북 표본점: {len(stand_rows)}")
    with open(STAND_OUT, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.DictWriter(f, fieldnames=stand_out_cols)
        w.writeheader()
        w.writerows(stand_rows)
    print(f"  저장: {STAND_OUT.name} ({STAND_OUT.stat().st_size // 1024} KB)")

    # ── 2. 입목조사표 ──
    print("\n[2] 입목조사표 추출 (충북 표본점만, 수고 cm->m 변환)...")
    ws2 = wb['입목조사표']
    header2 = [c.value for c in next(ws2.iter_rows(min_row=1, max_row=1))]
    idx2 = {name: i for i, name in enumerate(header2)}

    tree_cols = [c for c in TREE_MAP.keys() if c in idx2]
    # 최종 출력 컬럼: _수고_cm → 수고_m, _추정수고_cm → 추정수고_m 로 치환
    tree_out_cols = []
    for c in tree_cols:
        name = TREE_MAP[c]
        if name == '_수고_cm':
            tree_out_cols.append('수고_m')
        elif name == '_추정수고_cm':
            tree_out_cols.append('추정수고_m')
        else:
            tree_out_cols.append(name)

    tree_rows = []
    for r in ws2.iter_rows(min_row=2, values_only=True):
        sp_id = str(r[idx2['SP_ID']]) if 'SP_ID' in idx2 else None
        if sp_id in chungbuk_ids:
            row_out = {}
            for eng in tree_cols:
                name = TREE_MAP[eng]
                val = r[idx2[eng]] if eng in idx2 else None
                if name == '_수고_cm':
                    row_out['수고_m'] = cm_to_m(val)
                elif name == '_추정수고_cm':
                    row_out['추정수고_m'] = cm_to_m(val)
                else:
                    row_out[name] = val
            tree_rows.append(row_out)

    print(f"  충북 그루: {len(tree_rows)}")
    with open(TREE_OUT, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.DictWriter(f, fieldnames=tree_out_cols)
        w.writeheader()
        w.writerows(tree_rows)
    print(f"  저장: {TREE_OUT.name} ({TREE_OUT.stat().st_size // 1024} KB)")

    # ── 3. 검증 ──
    print(f"\n{'=' * 75}")
    print("검증:")
    print(f"{'=' * 75}")
    years = {}
    for s in stand_rows:
        y = s['조사연도']
        years[y] = years.get(y, 0) + 1
    print(f"  측정연도 분포: {dict(sorted(years.items()))}")

    imsang = {}
    for s in stand_rows:
        im = s['임상']
        imsang[im] = imsang.get(im, 0) + 1
    print(f"  임상 분포: {imsang}")

    dbhs = [float(t['DBH_cm']) for t in tree_rows
            if t.get('DBH_cm') not in (None, '')]
    if dbhs:
        print(f"  DBH 범위: {min(dbhs):.1f} ~ {max(dbhs):.1f} cm (평균 {sum(dbhs)/len(dbhs):.1f})")

    heights = [t['수고_m'] for t in tree_rows if t.get('수고_m') is not None]
    print(f"  수고_m 유효: {len(heights)} / {len(tree_rows)} ({len(heights)/len(tree_rows)*100:.1f}%)")
    if heights:
        print(f"  수고_m 범위: {min(heights):.1f} ~ {max(heights):.1f} m (평균 {sum(heights)/len(heights):.1f})")

    vols = [float(t['추정간재적']) for t in tree_rows
            if t.get('추정간재적') not in (None, '')]
    if vols:
        print(f"  추정간재적 범위: {min(vols):.4f} ~ {max(vols):.4f} m³")

    print(f"\n  → 수고_m 컬럼 6/7차 호환. fit_correct.py 가 그대로 읽음.")
    print(f"  다음: fit_correct.py v7 재실행 (5차 통합)")


if __name__ == "__main__":
    main()