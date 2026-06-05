"""
nfi5_coord_probe.py — NFI 5차 충북 좌표 체계 진단.

목적:
  · 5차 충북 표본점 좌표 (Bessel TM) 범위 확인
  · 7차 충북 좌표 (EPSG:5181) 와 비교
  · 같은 표본점 (고정표본점) 있으면 좌표 변환 검증
  · Bessel → 현대 좌표 변환 EPSG 결정
"""
import openpyxl
import csv
from pathlib import Path

ROOT = Path(__file__).resolve().parents[3]
NFI_DIR = ROOT / "module_bd" / "data" / "raw" / "nfi"
NFI5 = NFI_DIR / "mdb_NFI_5_수정.xlsx"
STAND7 = NFI_DIR / "nfi7_chungbuk_stand.csv"


def main():
    print("=" * 75)
    print("NFI 5차 충북 좌표 체계 진단")
    print("=" * 75)

    # 1. 5차 충북 표본점 추출
    print("\n[1] 5차 임분조사표 로딩 (충북 필터)...")
    wb = openpyxl.load_workbook(NFI5, read_only=True)
    ws = wb['임분조사표']

    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    # 컬럼 인덱스
    idx = {name: i for i, name in enumerate(header)}
    print(f"  GYEAR idx={idx.get('GYEAR')}, SP_ID idx={idx.get('SP_ID')}")
    print(f"  NAME_SIDO idx={idx.get('NAME_SIDO')}, NAME_SGG idx={idx.get('NAME_SGG')}")
    print(f"  BESSEL_TM_Y idx={idx.get('GB_BESSEL_TM_Y')}, "
          f"BESSEL_TM_X idx={idx.get('GB_BESSEL_TM_X')}")

    chungbuk_5 = []
    sido_set = set()
    for r in ws.iter_rows(min_row=2, values_only=True):
        sido = r[idx['NAME_SIDO']]
        sido_set.add(sido)
        if sido and ('충청북도' in str(sido) or '충북' in str(sido)):
            chungbuk_5.append({
                'gyear': r[idx['GYEAR']],
                'sp_id': r[idx['SP_ID']],
                'sgg': r[idx['NAME_SGG']],
                'emd': r[idx['NAME_EMD']],
                'y': r[idx['GB_BESSEL_TM_Y']],
                'x': r[idx['GB_BESSEL_TM_X']],
            })

    print(f"\n  5차 전체 시도: {sorted(str(s) for s in sido_set if s)}")
    print(f"  5차 충북 표본점: {len(chungbuk_5)}")

    if chungbuk_5:
        print(f"\n  5차 충북 좌표 예시 (5개):")
        for c in chungbuk_5[:5]:
            print(f"    {c['sp_id']} ({c['sgg']} {c['emd']}, {c['gyear']}): "
                  f"Y={c['y']}, X={c['x']}")

        ys = [c['y'] for c in chungbuk_5 if c['y'] is not None]
        xs = [c['x'] for c in chungbuk_5 if c['x'] is not None]
        print(f"\n  5차 충북 Y 범위: {min(ys):.0f} ~ {max(ys):.0f}")
        print(f"  5차 충북 X 범위: {min(xs):.0f} ~ {max(xs):.0f}")

        # 측정연도 분포
        years = {}
        for c in chungbuk_5:
            y = c['gyear']
            years[y] = years.get(y, 0) + 1
        print(f"\n  5차 충북 측정연도 분포: {dict(sorted(years.items()))}")

    # 2. 7차 충북 좌표 범위
    print(f"\n[2] 7차 충북 좌표 범위 (비교용)...")
    with open(STAND7, encoding='utf-8-sig') as f:
        rows7 = list(csv.DictReader(f))
    ys7 = [float(r['좌표N']) for r in rows7 if r['좌표N']]
    xs7 = [float(r['좌표E']) for r in rows7 if r['좌표E']]
    print(f"  7차 충북 좌표N(Y) 범위: {min(ys7):.0f} ~ {max(ys7):.0f}")
    print(f"  7차 충북 좌표E(X) 범위: {min(xs7):.0f} ~ {max(xs7):.0f}")
    print(f"  7차 충북 좌표 예시:")
    for r in rows7[:3]:
        print(f"    {r['표본점번호']}: N={r['좌표N']}, E={r['좌표E']}")

    # 3. 비교 분석
    print(f"\n{'=' * 75}")
    print("좌표 체계 비교:")
    print(f"{'=' * 75}")
    if chungbuk_5:
        print(f"  5차 충북 Y: {min(ys):.0f}~{max(ys):.0f}, "
              f"7차 충북 Y: {min(ys7):.0f}~{max(ys7):.0f}")
        print(f"  5차 충북 X: {min(xs):.0f}~{max(xs):.0f}, "
              f"7차 충북 X: {min(xs7):.0f}~{max(xs7):.0f}")
        print(f"\n  → 자릿수·범위 비교로 좌표 체계 판단")
        print(f"    · 비슷하면 같은 TM (변환 불필요/단순)")
        print(f"    · 다르면 Bessel → Korea2000 변환 필요 (pyproj)")

    # 4. 고정표본점 검증 (SP_ID 또는 SPGID 매칭)
    print(f"\n[4] 고정표본점 가능성 (SP_ID 패턴 비교):")
    if chungbuk_5:
        ids5 = set(str(c['sp_id']) for c in chungbuk_5)
        ids7 = set(str(r['표본점번호']) for r in rows7)
        print(f"  5차 충북 SP_ID 예시: {list(ids5)[:5]}")
        print(f"  7차 충북 표본점번호 예시: {list(ids7)[:5]}")
        common = ids5 & ids7
        print(f"  직접 공통 ID: {len(common)}")
        print(f"  → ID 형식 다르면 좌표 기반 매칭 필요")


if __name__ == "__main__":
    main()