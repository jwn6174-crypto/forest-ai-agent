"""
nfi_5_6_7_probe.py — NFI 5·6·7차 구조 비교 진단.

목적:
  · 임분조사표 (3 차수) 컬럼 비교
  · 임목/입목조사표 (3 차수) 컬럼 비교
  · 측정 연도 컬럼 위치 + 좌표 컬럼 통일성
  · 충북 표본점 수 (3 차수)
  
→ 시계열 통합 가능성 진단
"""
from openpyxl import load_workbook
from pathlib import Path
from collections import Counter

NFI_DIR = Path("module_bd/data/raw/nfi")


def probe_sheet(wb, sheet_name, n_rows=5):
    """시트 헤더 + n_rows 데이터 행 추출."""
    if sheet_name not in wb.sheetnames:
        return None, None
    ws = wb[sheet_name]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= n_rows + 1:
            break
        rows.append(row)
    if not rows:
        return None, None
    header = rows[0]
    data = rows[1:]
    return header, data


def count_chungbuk(wb, sheet_name, sido_col_name):
    """충북 표본점 수 카운트."""
    if sheet_name not in wb.sheetnames:
        return 0
    ws = wb[sheet_name]
    # 헤더에서 sido 컬럼 인덱스
    header = next(ws.iter_rows(max_row=1, values_only=True))
    if sido_col_name not in header:
        return 0
    col_idx = header.index(sido_col_name)
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[col_idx] == '충청북도':
            count += 1
    return count


def main():
    print("=" * 80)
    print("NFI 5·6·7차 구조 비교 진단")
    print("=" * 80)

    sheet_targets = {
        5: ['임분조사표', '입목조사표', '임목조사표'],  # 5차 입목/임목 둘 다 시도
        6: ['임분조사표', '임목조사표'],
        7: ['임분조사표', '임목조사표'],
    }

    for v in [5, 6, 7]:
        p = NFI_DIR / f"mdb_NFI_{v}_수정.xlsx"
        if not p.exists():
            print(f"\n⚠ NFI {v}: 파일 없음")
            continue

        print(f"\n{'=' * 80}")
        print(f"NFI {v}차: {p.name}")
        print(f"{'=' * 80}")
        wb = load_workbook(p, read_only=True)
        print(f"  시트 목록: {wb.sheetnames}")

        for sheet in sheet_targets[v]:
            if sheet not in wb.sheetnames:
                continue
            print(f"\n  [시트: {sheet}]")
            header, data = probe_sheet(wb, sheet, n_rows=2)
            if header is None:
                continue
            print(f"    컬럼 수: {len(header)}")
            print(f"    컬럼: {list(header)}")
            if data:
                print(f"    예시 1행: {data[0][:10]}...")  # 처음 10 컬럼만


if __name__ == "__main__":
    main()