"""
nfi6_extract.py — NFI 6차 충북 표본점·나무 추출.

목적:
  · NFI 6차 임분조사표 → 충북 stand6.csv
  · NFI 6차 임목조사표 → 충북 tree6.csv  
  · 7차 nfi_extract.py 패턴 그대로 (6차 컬럼 호환 확인됨)

D11 단위 변환 적용:
  · 수고/지하고 cm → m
  · DBH cm 그대로
  
출력:
  module_bd/data/raw/nfi/nfi6_chungbuk_stand.csv
  module_bd/data/raw/nfi/nfi6_chungbuk_tree.csv
"""
from openpyxl import load_workbook
from pathlib import Path
import csv

ROOT = Path(__file__).resolve().parents[3]
NFI_PATH = ROOT / "module_bd" / "data" / "raw" / "nfi" / "mdb_NFI_6_수정.xlsx"
OUT_DIR = ROOT / "module_bd" / "data" / "raw" / "nfi"

TARGET_SIDO = "충청북도"


def safe_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def cm_to_m(v):
    f = safe_float(v)
    return f / 100 if f is not None else None


def extract_stand(wb):
    """6차 임분조사표 추출."""
    print(f"6차 임분조사표 로드 중...")
    ws = wb["임분조사표"]
    header = next(ws.iter_rows(max_row=1, values_only=True))
    col_idx = {col: i for i, col in enumerate(header)}

    cols = [
        '집락번호', '표본점번호', '조사연도', '조사차기',
        '광역시도', '시군구', '읍면동',
        '좌표N', '좌표E',
        '해발고', '경사', '8방위코드',
        '지형', '사면위치',
        '임상', '수관밀도', '경급', '영급',
        '소유', '임종', '지종', '갱신형태',
        '토양형', '토성(A)', '토성(B)',
    ]

    out = []
    total = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        total += 1
        sido = row[col_idx.get('광역시도', -1)] if '광역시도' in col_idx else None
        if sido != TARGET_SIDO:
            continue
        record = {}
        for c in cols:
            if c in col_idx:
                v = row[col_idx[c]]
                if c in ('해발고', '경사'):
                    record[c] = safe_float(v)
                else:
                    record[c] = v
        out.append(record)
    print(f"  전국 표본점: {total:,}")
    print(f"  충북 표본점: {len(out)}")
    return out


def extract_tree(wb, plot_ids):
    """6차 임목조사표 추출 (단위 변환)."""
    print(f"\n6차 임목조사표 로드 중...")
    ws = wb["임목조사표"]
    header = next(ws.iter_rows(max_row=1, values_only=True))
    col_idx = {col: i for i, col in enumerate(header)}

    cols_raw = [
        '집락번호', '표본점번호', '번호', '조사차기',
        '학명번호', '수종명', '교목구분', '침활구분',
        '흉고직경',
        '형질급', '수관급', '수관활력도',
        '지하고', '수고',
        '거리(m)', '방위각(º)',
        '추정수고', '추정간재적', '표준목간재적',
        '비율',
    ]

    # 단면적 컬럼명: 7차 = '단면적', 6차 = '흉고단면적'
    if '흉고단면적' in col_idx:
        cols_raw.append('흉고단면적')
    elif '단면적' in col_idx:
        cols_raw.append('단면적')

    out = []
    total = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        total += 1
        plot_id = row[col_idx.get('표본점번호', -1)]
        if plot_id not in plot_ids:
            continue
        record = {}
        for c in cols_raw:
            if c in col_idx:
                v = row[col_idx[c]]
                if c in ('수고', '지하고', '추정수고'):
                    record[c + '_m'] = cm_to_m(v)
                elif c == '흉고직경':
                    record['DBH_cm'] = safe_float(v)
                elif c in ('흉고단면적', '단면적', '추정간재적', '표준목간재적'):
                    record[c] = safe_float(v)
                else:
                    record[c] = v
        out.append(record)
    print(f"  전국 나무: {total:,}")
    print(f"  충북 나무: {len(out):,}")
    return out


def write_csv(records, path):
    if not records:
        print(f"  ⚠ 데이터 없음: {path.name}")
        return
    keys = list(records[0].keys())
    with open(path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=keys)
        writer.writeheader()
        writer.writerows(records)
    size_kb = path.stat().st_size / 1024
    print(f"  ✓ {path.name}: {len(records)}행, {size_kb:.0f} KB")


def main():
    print("=" * 70)
    print("NFI 6차 충북 추출 (7차 패턴 재사용)")
    print("=" * 70)

    if not NFI_PATH.exists():
        print(f"⚠ 파일 없음: {NFI_PATH}")
        return

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    print(f"파일 로드: {NFI_PATH.name} (127MB, 1-2분)")
    wb = load_workbook(NFI_PATH, read_only=True)

    stand_records = extract_stand(wb)
    plot_ids = {r['표본점번호'] for r in stand_records}
    tree_records = extract_tree(wb, plot_ids)

    print(f"\nCSV 저장:")
    stand_path = OUT_DIR / "nfi6_chungbuk_stand.csv"
    tree_path = OUT_DIR / "nfi6_chungbuk_tree.csv"
    write_csv(stand_records, stand_path)
    write_csv(tree_records, tree_path)

    # 보은 매칭 검증
    boeun = [r for r in stand_records if r.get('시군구') == '보은군']
    boeun_ids = {r['표본점번호'] for r in boeun}
    boeun_trees = [t for t in tree_records if t.get('표본점번호') in boeun_ids]

    print(f"\n{'=' * 70}")
    print(f"6차 충북 추출 결과:")
    print(f"  충북: {len(stand_records)} 표본점, {len(tree_records):,} 나무")
    print(f"  보은: {len(boeun)} 표본점, {len(boeun_trees):,} 나무")
    print(f"\n  7차 비교 참고:")
    print(f"    7차 충북: 1162 / 46,722")
    print(f"    7차 보은: 102 / 4,505")
    print(f"{'=' * 70}")


if __name__ == "__main__":
    main()