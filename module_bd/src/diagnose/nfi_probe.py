"""
nfi_probe.py — NFI 7차 보은 표본점 추출 진단.

목적:
  · NFI 7차 임분조사표 + 임목조사표에서 보은 데이터 추출·진단
  · climate_correct·Weibull 분석에 *충분한 데이터인가* 확인

데이터 출처: 국립산림과학원 NFI 7차 (2016~2020 조사)
보안: raw xlsx 는 git 차단. 추출본 csv 만 추후 git 가능.

진단 결과 (2026-05-27):
  · 임분조사표: 전국 16,617개 / 충북 1,162개 / 보은 102개
  · 가이드 §3.4 추정 27~30개의 3.6배. 통계 신뢰성 강화.
  · 해발고 96~678m, 산악기상 관측소(242~627m) 와 매칭 가능.
"""
from openpyxl import load_workbook
from pathlib import Path
from collections import Counter

ROOT = Path(__file__).resolve().parents[3]
NFI_PATH = ROOT / "module_bd" / "data" / "raw" / "nfi" / "mdb_NFI_7_수정.xlsx"


def safe_numeric(values):
    """str·int 섞인 리스트 → float 만 추출."""
    out = []
    for v in values:
        try:
            out.append(float(v))
        except (ValueError, TypeError):
            pass
    return out


def find_chungbuk_boeun(ws):
    """충북 + 보은 표본점 추출 (임분조사표)."""
    header = next(ws.iter_rows(max_row=1, values_only=True))
    col_idx = {col: i for i, col in enumerate(header)}

    needed_cols = ['집락번호', '표본점번호', '광역시도', '시군구', '읍면동',
                   '좌표N', '좌표E', '해발고', '경사', '임상', '영급', '경급',
                   '소유', '임종']
    missing = [c for c in needed_cols if c not in col_idx]
    if missing:
        print(f"  ⚠ 헤더에 없는 칼럼: {missing}")

    chungbuk_plots = []
    boeun_plots = []
    total = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        total += 1
        sido = row[col_idx['광역시도']]
        if sido != '충청북도':
            continue

        plot = {c: row[col_idx[c]] for c in needed_cols if c in col_idx}
        chungbuk_plots.append(plot)
        if plot.get('시군구') == '보은군':
            boeun_plots.append(plot)

    return chungbuk_plots, boeun_plots, total


def summarize_plots(plots, label):
    """추출 표본점 요약 (시군 분포)."""
    print(f"\n[{label}] {len(plots)}개 표본점")
    if not plots:
        return
    sigungu_counts = Counter(p.get('시군구') for p in plots)
    print(f"  시군구 분포 ({len(sigungu_counts)}개 시군):")
    for sg, n in sigungu_counts.most_common():
        print(f"    {sg}: {n}개")


def summarize_boeun_detail(plots):
    """보은 표본점 상세 분포."""
    if not plots:
        print("  ⚠ 보은 표본점 0개")
        return

    print(f"\n[보은군 표본점 분포]")
    for label, key in [('임상', '임상'), ('영급', '영급'), ('읍면동', '읍면동')]:
        counts = Counter(p.get(key) for p in plots)
        print(f"  {label}별 ({len(counts)}개):")
        for k, n in counts.most_common():
            print(f"    {k or '(결측)'}: {n}개")

    elevs = safe_numeric([p.get('해발고') for p in plots
                          if p.get('해발고') is not None])
    if elevs:
        print(f"  해발고: {min(elevs):.0f} ~ {max(elevs):.0f}m "
              f"(평균 {sum(elevs)/len(elevs):.0f}m, 유효 {len(elevs)}/{len(plots)})")

    slopes = safe_numeric([p.get('경사') for p in plots
                           if p.get('경사') is not None])
    if slopes:
        print(f"  경사: {min(slopes):.0f} ~ {max(slopes):.0f}° "
              f"(평균 {sum(slopes)/len(slopes):.0f}°)")


def find_boeun_trees(wb, boeun_plot_ids):
    """임목조사표에서 보은 표본점의 개별 나무 추출.

    Args:
        wb: 워크북
        boeun_plot_ids: 보은 표본점번호 set
    """
    print(f"\n임목조사표 시트 로드 중...")
    ws = wb["임목조사표"]
    header = next(ws.iter_rows(max_row=1, values_only=True))
    col_idx = {col: i for i, col in enumerate(header)}

    needed = ['집락번호', '표본점번호', '학명번호', '수종명', '교목구분',
              '침활구분', '흉고직경', '수고', '수령', '추정간재적',
              '형질급', '수관급']

    trees = []
    total = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        total += 1
        plot_id = row[col_idx['표본점번호']]
        if plot_id not in boeun_plot_ids:
            continue
        tree = {c: row[col_idx[c]] for c in needed if c in col_idx}
        trees.append(tree)

    print(f"  전국 개별 나무: {total:,}그루")
    print(f"  보은 개별 나무: {len(trees)}그루")
    return trees


def summarize_boeun_trees(trees):
    """보은 개별 나무 분포 — Weibull fit input 확인."""
    if not trees:
        print("  ⚠ 보은 나무 0그루")
        return

    print(f"\n[보은 개별 나무 분포 — Weibull fit input]")

    # 수종 분포 (상위 15개)
    species_counts = Counter(t.get('수종명') for t in trees)
    print(f"  수종별 ({len(species_counts)}종, 상위 15개):")
    for sp, n in species_counts.most_common(15):
        print(f"    {sp or '(결측)'}: {n}그루")

    # 침활구분
    chimhwal_counts = Counter(t.get('침활구분') for t in trees)
    print(f"  침활구분:")
    for k, n in chimhwal_counts.most_common():
        print(f"    {k or '(결측)'}: {n}그루")

    # DBH 분포 (Weibull fit 의 input)
    dbh = safe_numeric([t.get('흉고직경') for t in trees
                        if t.get('흉고직경') is not None])
    if dbh:
        print(f"  흉고직경(DBH): {min(dbh):.1f} ~ {max(dbh):.1f}cm "
              f"(평균 {sum(dbh)/len(dbh):.1f}cm, 유효 {len(dbh)}/{len(trees)})")
        # 등급 분포 — 가이드 §2.4 등급 기준 (소경·중경·대경)
        small = sum(1 for d in dbh if d < 18)
        medium = sum(1 for d in dbh if 18 <= d < 30)
        large = sum(1 for d in dbh if d >= 30)
        print(f"    소경(<18cm):   {small}그루 ({small/len(dbh)*100:.0f}%)")
        print(f"    중경(18-30cm): {medium}그루 ({medium/len(dbh)*100:.0f}%)")
        print(f"    대경(≥30cm):   {large}그루 ({large/len(dbh)*100:.0f}%)")

    # 수고
    height = safe_numeric([t.get('수고') for t in trees
                           if t.get('수고') is not None])
    if height:
        print(f"  수고: {min(height):.1f} ~ {max(height):.1f}m "
              f"(평균 {sum(height)/len(height):.1f}m, 유효 {len(height)}/{len(trees)})")

    # 수령
    age = safe_numeric([t.get('수령') for t in trees
                        if t.get('수령') is not None])
    if age:
        print(f"  수령: {min(age):.0f} ~ {max(age):.0f}년 "
              f"(평균 {sum(age)/len(age):.0f}년, 유효 {len(age)}/{len(trees)})")

    # 추정간재적
    vol = safe_numeric([t.get('추정간재적') for t in trees
                        if t.get('추정간재적') is not None])
    if vol:
        print(f"  추정간재적: {min(vol):.3f} ~ {max(vol):.3f}m³ "
              f"(평균 {sum(vol)/len(vol):.3f}m³, 유효 {len(vol)}/{len(trees)})")


def main():
    print("=" * 60)
    print("NFI 7차 — 충북·보은 표본점·나무 진단")
    print("=" * 60)

    if not NFI_PATH.exists():
        print(f"⚠ 파일 없음: {NFI_PATH}")
        return

    # 임분조사표 (표본점 단위)
    print(f"파일 로드 중: {NFI_PATH.name} (124MB, 1~2분 소요)")
    wb = load_workbook(NFI_PATH, read_only=True)
    ws_imbun = wb["임분조사표"]
    chungbuk, boeun, total = find_chungbuk_boeun(ws_imbun)

    print(f"\n전국 표본점 (조사 가능 전체): {total:,}개")
    summarize_plots(chungbuk, "충청북도")
    summarize_plots(boeun, "보은군")
    summarize_boeun_detail(boeun)

    # 보은 표본점 ID set 만들기
    boeun_ids = {p['표본점번호'] for p in boeun}

    # 임목조사표 (개별 나무 단위)
    trees = find_boeun_trees(wb, boeun_ids)
    summarize_boeun_trees(trees)

    # 가이드 비교
    print(f"\n{'=' * 60}")
    print(f"가이드 §3.4 vs 실측:")
    print(f"  보은 표본점: 추정 27~30 → 실측 {len(boeun)}개 (3.6배)")
    print(f"  보은 개별 나무: {len(trees)}그루")
    if len(trees) >= 1000:
        print(f"  → Weibull fit *수종·영급별 sub-fit* 충분히 가능.")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()