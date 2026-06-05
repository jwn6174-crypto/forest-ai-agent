"""
nfi_extract.py — NFI 7차 충북 표본점·나무 추출.

목적:
  · NFI 7차 임분조사표 → 충북 1,162 표본점 → stand.csv
  · NFI 7차 임목조사표 → 충북 약 5만 나무 → tree.csv
  · 단위 변환 적용 (D11 표준): 수고·지하고 cm → m
  · 매번 124MB xlsx 로딩 회피 (csv 약 5MB, 0.1초 로딩)

데이터 출처: 국립산림과학원 NFI 7차 (2016~2020)
관련 결정: D11 (단위·코드·구조), D12 (csv 저장 방식, 옵션 A — 좌표 그대로)
보안: raw/nfi/ 안에 저장. gitignore 차단. *집계 결과만 git*.

실행 시간: 약 3~5분 (xlsx 두 시트 로드)
출력:
  module_bd/data/raw/nfi/nfi7_chungbuk_stand.csv  (~200KB)
  module_bd/data/raw/nfi/nfi7_chungbuk_tree.csv   (~5MB)
"""
from openpyxl import load_workbook
from pathlib import Path
import csv

ROOT = Path(__file__).resolve().parents[3]
NFI_PATH = ROOT / "module_bd" / "data" / "raw" / "nfi" / "mdb_NFI_7_수정.xlsx"
OUT_DIR = ROOT / "module_bd" / "data" / "raw" / "nfi"

TARGET_SIDO = "충청북도"


def safe_float(v):
    """str·int·None 섞인 값을 float 로 안전 변환. 실패 시 None."""
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def cm_to_m(v):
    """NFI cm 값 → m 변환 (D11 결정 5)."""
    f = safe_float(v)
    return f / 100 if f is not None else None


def extract_stand(wb):
    """임분조사표에서 충북 표본점 추출."""
    print(f"임분조사표 로드 중...")
    ws = wb["임분조사표"]
    header = next(ws.iter_rows(max_row=1, values_only=True))
    col_idx = {col: i for i, col in enumerate(header)}

    # 추출 칼럼들 (D11 단위·코드 표 기반)
    cols = [
        '집락번호', '표본점번호', '조사연도',
        '광역시도', '시군구', '읍면동',
        '좌표N', '좌표E',
        '해발고', '경사', '8방위코드',
        '지형', '사면위치',
        '임상', '수관밀도', '경급', '영급',
        '소유', '임종', '지종', '갱신형태',
        '토양형', '토성(A)', '토성(B)',
    ]

    out = []
    total = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        total += 1
        sido = row[col_idx['광역시도']]
        if sido != TARGET_SIDO:
            continue

        record = {}
        for c in cols:
            if c in col_idx:
                v = row[col_idx[c]]
                # 해발고·경사는 숫자로 변환
                if c in ('해발고', '경사'):
                    record[c] = safe_float(v)
                else:
                    record[c] = v
        out.append(record)

    print(f"  전국 표본점: {total:,}개")
    print(f"  충북 표본점: {len(out)}개")
    return out, cols


def extract_tree(wb, plot_ids):
    """임목조사표에서 충북 표본점의 개별 나무 추출."""
    print(f"\n임목조사표 로드 중 (시간 더 걸림)...")
    ws = wb["임목조사표"]
    header = next(ws.iter_rows(max_row=1, values_only=True))
    col_idx = {col: i for i, col in enumerate(header)}

    # 추출 칼럼들 (D11 결정 5 단위 변환 적용)
    cols_raw = [
        '집락번호', '표본점번호', '번호',
        '학명번호', '수종명', '교목구분', '침활구분',
        '흉고직경', '단면적',
        '형질급', '수관급', '수관활력도',
        '지하고', '수고',  # cm → m 변환 예정
        '거리', '방위각',
        '추정수고', '추정간재적', '표준목간재적',
        '비율',
    ]

    out = []
    total = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        total += 1
        plot_id = row[col_idx['표본점번호']]
        if plot_id not in plot_ids:
            continue

        record = {}
        for c in cols_raw:
            if c in col_idx:
                v = row[col_idx[c]]
                # 단위 변환 (D11 결정 5)
                if c in ('수고', '지하고', '추정수고'):
                    record[c + '_m'] = cm_to_m(v)
                elif c == '흉고직경':
                    record['DBH_cm'] = safe_float(v)
                elif c in ('단면적', '추정간재적', '표준목간재적'):
                    record[c] = safe_float(v)
                else:
                    record[c] = v
        out.append(record)

    print(f"  전국 나무: {total:,}그루")
    print(f"  충북 나무: {len(out)}그루")
    return out


def write_csv(records, cols, path):
    """list[dict] → csv 저장 (UTF-8 BOM, Excel 호환)."""
    if not records:
        print(f"  ⚠ 데이터 없음, 저장 건너뜀: {path.name}")
        return

    # 실제 키 추출 (변환 컬럼명 반영)
    sample_keys = list(records[0].keys())

    with open(path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=sample_keys)
        writer.writeheader()
        writer.writerows(records)

    size_kb = path.stat().st_size / 1024
    print(f"  ✓ {path.name}: {len(records)}행, {size_kb:.0f} KB")


def main():
    print("=" * 60)
    print(f"NFI 7차 충북 표본점·나무 추출 (D11·D12 적용)")
    print("=" * 60)

    if not NFI_PATH.exists():
        print(f"⚠ 파일 없음: {NFI_PATH}")
        return

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    print(f"파일 로드: {NFI_PATH.name} (124MB, 1~2분)")
    wb = load_workbook(NFI_PATH, read_only=True)

    # 1. 표본점
    stand_records, stand_cols = extract_stand(wb)
    plot_ids = {r['표본점번호'] for r in stand_records}

    # 2. 나무
    tree_records = extract_tree(wb, plot_ids)

    # 3. 저장
    print(f"\nCSV 저장:")
    stand_path = OUT_DIR / "nfi7_chungbuk_stand.csv"
    tree_path = OUT_DIR / "nfi7_chungbuk_tree.csv"
    write_csv(stand_records, stand_cols, stand_path)
    write_csv(tree_records, [], tree_path)  # tree 는 키 자동 추출

    # 4. 요약
    boeun = [r for r in stand_records if r.get('시군구') == '보은군']
    boeun_ids = {r['표본점번호'] for r in boeun}
    boeun_trees = [t for t in tree_records if t.get('표본점번호') in boeun_ids]

    print(f"\n{'=' * 60}")
    print(f"추출 결과 요약:")
    print(f"  충북 1162 → 보은 {len(boeun)} (실측 102 매칭)")
    print(f"  충북 나무 {len(tree_records):,} → 보은 {len(boeun_trees):,}")
    print(f"\n다음 단계 (D13·D14):")
    print(f"  · D13 climate_correct: stand.csv + mt_weather csv 결합")
    print(f"  · D14 Weibull fit: tree.csv 수종·영급별 DBH 분포 fit")
    print(f"{'=' * 60}")
    print(f"\n보안: 두 csv 는 raw/nfi/ 안. gitignore 차단 유지.")


if __name__ == "__main__":
    main()